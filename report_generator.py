"""
Generates a polished multi-sheet Excel report from portfolio analysis results:
Holdings (with live formulas), Summary (KPIs + sector allocation), Charts.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="808080")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
CURRENCY_FMT = '"₹"#,##0.00'
PCT_FMT = '0.00"%"'
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
RED_FILL = PatternFill("solid", start_color="FFC7CE")


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def autofit_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 3, max_width))


def build_holdings_sheet(wb, result):
    ws = wb.active
    ws.title = "Holdings"

    headers = ["Symbol", "Company", "Sector", "Qty", "Buy Price", "Current Price",
               "Invested", "Current Value", "Gain/Loss", "Gain/Loss %", "Weight %"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    r = 2
    for h in result["holdings"]:
        ws.cell(r, 1, h["symbol"])
        ws.cell(r, 2, h["company_name"])
        ws.cell(r, 3, h["sector"])
        ws.cell(r, 4, h["quantity"])
        ws.cell(r, 5, h["buy_price"])
        ws.cell(r, 6, h["current_price"])
        ws.cell(r, 7, f"=D{r}*E{r}")
        ws.cell(r, 8, f"=D{r}*F{r}")
        ws.cell(r, 9, f"=H{r}-G{r}")
        ws.cell(r, 10, f"=IF(G{r}=0,0,(I{r}/G{r})*100)")
        ws.cell(r, 11, f"=IF($H${len(result['holdings'])+1}=0,0,(H{r}/$H${len(result['holdings'])+1})*100)")
        r += 1
    last_row = r - 1

    total_row = r
    ws.cell(total_row, 2, "TOTAL").font = BOLD_FONT
    ws.cell(total_row, 7, f"=SUM(G2:G{last_row})")
    ws.cell(total_row, 8, f"=SUM(H2:H{last_row})")
    ws.cell(total_row, 9, f"=H{total_row}-G{total_row}")
    ws.cell(total_row, 10, f"=IF(G{total_row}=0,0,(I{total_row}/G{total_row})*100)")
    for c in range(1, len(headers) + 1):
        ws.cell(total_row, c).font = BOLD_FONT
        ws.cell(total_row, c).fill = PatternFill("solid", start_color="D9E1F2")

    for row in range(2, total_row + 1):
        for col in (5, 6, 7, 8, 9):
            ws.cell(row, col).number_format = CURRENCY_FMT
        ws.cell(row, 10).number_format = PCT_FMT
        ws.cell(row, 11).number_format = PCT_FMT
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).border = THIN_BORDER
            if ws.cell(row, col).font != BOLD_FONT:
                ws.cell(row, col).font = NORMAL_FONT

    ws.conditional_formatting.add(f"I2:I{last_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL))
    ws.conditional_formatting.add(f"I2:I{last_row}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{last_row}"
    autofit_columns(ws)
    return ws, last_row, total_row


def build_summary_sheet(wb, result, holdings_last_row, holdings_total_row):
    ws = wb.create_sheet("Summary")

    ws["A1"] = "Portfolio Performance Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")

    kpis = [
        ("Total Invested", f"=Holdings!G{holdings_total_row}", CURRENCY_FMT),
        ("Current Value", f"=Holdings!H{holdings_total_row}", CURRENCY_FMT),
        ("Total Gain/Loss", f"=Holdings!I{holdings_total_row}", CURRENCY_FMT),
        ("Return %", f"=Holdings!J{holdings_total_row}", PCT_FMT),
    ]
    for i, (label, formula, fmt) in enumerate(kpis, start=1):
        c = ws.cell(4, i, label)
        c.font = BOLD_FONT
        c.fill = PatternFill("solid", start_color="D9E1F2")
        v = ws.cell(5, i, formula)
        v.font = Font(name="Arial", bold=True, size=13, color="1F4E78")
        v.number_format = fmt

    ws["A8"] = "Sector Allocation"
    ws["A8"].font = BOLD_FONT
    header_row = 9
    ws.cell(header_row, 1, "Sector")
    ws.cell(header_row, 2, "Value")
    style_header_row(ws, header_row, 2)
    r = header_row + 1
    sector_start = r
    for sector, value in sorted(result["sector_allocation"].items(), key=lambda x: -x[1]):
        ws.cell(r, 1, sector).font = NORMAL_FONT
        ws.cell(r, 2, value).number_format = CURRENCY_FMT
        for c in range(1, 3):
            ws.cell(r, c).border = THIN_BORDER
        r += 1
    sector_end = r - 1

    bw_row = sector_end + 3
    ws.cell(bw_row, 1, "Best Performer").font = BOLD_FONT
    ws.cell(bw_row, 2, f"{result['best_performer']['symbol']} ({result['best_performer']['gain_loss_pct']}%)")
    ws.cell(bw_row + 1, 1, "Worst Performer").font = BOLD_FONT
    ws.cell(bw_row + 1, 2, f"{result['worst_performer']['symbol']} ({result['worst_performer']['gain_loss_pct']}%)")

    autofit_columns(ws)
    return ws, {"sector_range": (sector_start, sector_end)}


def build_charts_sheet(wb, holdings_ws, summary_ws, holdings_last_row, ranges):
    ws = wb.create_sheet("Charts")
    ws["A1"] = "Visual Insights"
    ws["A1"].font = TITLE_FONT

    bar = BarChart()
    bar.title = "Gain/Loss by Stock"
    bar.y_axis.title = "Gain/Loss (₹)"
    data = Reference(holdings_ws, min_col=9, min_row=1, max_row=holdings_last_row)
    cats = Reference(holdings_ws, min_col=1, min_row=2, max_row=holdings_last_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height, bar.width = 9, 16
    ws.add_chart(bar, "A3")

    s0, s1 = ranges["sector_range"]
    pie = PieChart()
    pie.title = "Portfolio Allocation by Sector"
    data2 = Reference(summary_ws, min_col=2, min_row=s0 - 1, max_row=s1)
    cats2 = Reference(summary_ws, min_col=1, min_row=s0, max_row=s1)
    pie.add_data(data2, titles_from_data=True)
    pie.set_categories(cats2)
    pie.height, pie.width = 9, 16
    ws.add_chart(pie, "A22")

    return ws


def generate_report(result, output_path=None):
    if output_path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"Portfolio_Report_{stamp}.xlsx"

    wb = Workbook()
    holdings_ws, last_row, total_row = build_holdings_sheet(wb, result)
    summary_ws, ranges = build_summary_sheet(wb, result, last_row, total_row)
    build_charts_sheet(wb, holdings_ws, summary_ws, last_row, ranges)

    wb.move_sheet("Summary", offset=-1)
    wb.active = 0
    wb.save(output_path)
    return output_path